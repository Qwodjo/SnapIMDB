import os
import io
import json
from collections import defaultdict

from fastapi import FastAPI, File, UploadFile, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from dotenv import load_dotenv

from app.schemas import FIELD_KEYS, EXCEL_COLUMN_NAMES
from app.extraction import extract_from_multiple_images
from app.fusion import fuse_multiple_images, group_images_by_product
from app.critic import run_critic_pass
from app.normalization import normalize_record, normalize_field_value
from app.duplicates import run_duplicate_check
from app.export import build_predictions_excel

load_dotenv()

app = FastAPI(
    title="SnapIMDB API",
    description="AI-Driven Image-to-Item Master Data Tool",
    version="1.0.0",
)

# FIXED PRODUCTION-SAFE CORS SETUP:
# This single block securely handles your local machine AND any Vercel domain.
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"(http://localhost:5173|https://.*\.vercel\.app)",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Health 

@app.get("/health")
def health():
    return {
        "status": "ok",
        "gemini_key_set": bool(os.getenv("GEMINI_API_KEY")),
    }


# Single product extraction 

@app.post("/extract")
async def extract_product(files: list[UploadFile] = File(...)):
    """
    Accept 1–8 images of the SAME product.
    Runs parallel extraction, confidence-weighted fusion,
    critic pass, and normalization.
    Returns one IMDBRecord dict.
    """
    if not files:
        raise HTTPException(400, "No files uploaded")
    if len(files) > 8:
        raise HTTPException(400, "Maximum 8 images per product")

    # Read all files
    images = []
    for f in files:
        content = await f.read()
        images.append((content, f.content_type or "image/jpeg", f.filename or "image.jpg"))

    # Step 1: Extract from all images in parallel
    raw_results = await extract_from_multiple_images(images)

    # Step 2: Fuse results by confidence
    fused = fuse_multiple_images(raw_results)

    # Step 3: Critic pass — self-review for inconsistencies
    reviewed = await run_critic_pass(fused)

    # Step 4: Final normalization
    from app.schemas import IMDBRecord, IMDBField
    def dict_to_record(d):
        kwargs = {}
        for field in FIELD_KEYS:
            raw = d.get(field, {})
            if isinstance(raw, dict):
                kwargs[field] = IMDBField(
                    value=raw.get("value", ""),
                    confidence=float(raw.get("confidence", 0.0)),
                )
            else:
                kwargs[field] = IMDBField()
        kwargs["image_id"] = d.get("image_id", "")
        kwargs["has_conflicts"] = d.get("has_conflicts", False)
        kwargs["conflict_notes"] = d.get("conflict_notes", "")
        return IMDBRecord(**kwargs)

    record = dict_to_record(reviewed)
    record = normalize_record(record)

    return record.model_dump()


# Batch extraction 

@app.post("/batch-extract")
async def batch_extract(files: list[UploadFile] = File(...)):
    """
    Accept images for MULTIPLE products.
    Groups by filename prefix automatically.
    Runs the full pipeline per product group.
    Returns all records with duplicate flags.
    """
    if not files:
        raise HTTPException(400, "No files uploaded")

    # Read all files into memory
    file_data = []
    filenames = []
    for f in files:
        content = await f.read()
        fname = f.filename or f"image_{len(file_data)}.jpg"
        file_data.append((content, f.content_type or "image/jpeg", fname))
        filenames.append(fname)

    # Group by product prefix
    groups = group_images_by_product(filenames)

    all_records = []

    for prefix, indices in groups.items():
        group_images = [file_data[i] for i in indices]

        # Step 1: Extract
        raw_results = await extract_from_multiple_images(group_images)

        # Step 2: Fuse
        fused = fuse_multiple_images(raw_results)
        fused["image_id"] = prefix

        # Step 3: Critic
        reviewed = await run_critic_pass(fused)

        # Step 4: Normalize
        from app.schemas import IMDBRecord, IMDBField
        def dict_to_record(d):
            kwargs = {}
            for field in FIELD_KEYS:
                raw = d.get(field, {})
                if isinstance(raw, dict):
                    kwargs[field] = IMDBField(
                        value=raw.get("value", ""),
                        confidence=float(raw.get("confidence", 0.0)),
                    )
                else:
                    kwargs[field] = IMDBField()
            kwargs["image_id"] = d.get("image_id", "")
            kwargs["has_conflicts"] = d.get("has_conflicts", False)
            kwargs["conflict_notes"] = d.get("conflict_notes", "")
            return IMDBRecord(**kwargs)

        record = dict_to_record(reviewed)
        record = normalize_record(record)
        all_records.append(record.model_dump())

    # Step 5: Duplicate detection across all products
    all_records = run_duplicate_check(all_records)

    return {
        "records": all_records,
        "total": len(all_records),
        "duplicates_found": sum(1 for r in all_records if r.get("is_duplicate", False)),
        "conflicts_found": sum(1 for r in all_records if r.get("has_conflicts", False)),
    }


# Export 

@app.post("/export")
async def export_excel(records: list[dict] = Body(...)):
    """
    Accept a list of IMDBRecord dicts.
    Returns a predictions.xlsx file as a binary download.
    """
    if not records:
        raise HTTPException(400, "No records to export")

    excel_bytes = build_predictions_excel(records)

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=predictions.xlsx"},
    )


#  Normalize single field 

@app.post("/normalize-field")
def normalize_single_field(payload: dict = Body(...)):
    """
    Normalize a single edited field value from the frontend.
    Used when user edits a cell inline.
    Body: { "field": "packaging_type", "value": "plstc btl" }
    Returns: { "normalized": "PLASTIC BOTTLE" }
    """
    field = payload.get("field", "")
    value = payload.get("value", "")
    if not field or not value:
        raise HTTPException(400, "field and value are required")
    normalized = normalize_field_value(field, value)
    return {"normalized": normalized}


# Evaluate vs ground truth 

@app.post("/evaluate")
async def evaluate(
    predictions: list[dict] = Body(..., embed=True),
    ground_truth: list[dict] = Body(..., embed=True),
):
    """
    Compare prediction records against ground truth records.
    Matches by position (index order).
    Returns per-column accuracy + overall score.
    """
    if not predictions or not ground_truth:
        raise HTTPException(400, "Both predictions and ground_truth are required")

    n = min(len(predictions), len(ground_truth))

    totals  = {k: 0 for k in FIELD_KEYS}
    matches = {k: 0 for k in FIELD_KEYS}

    for i in range(n):
        pred = predictions[i]
        gt   = ground_truth[i]

        for key in FIELD_KEYS:
            # Get predicted value
            pred_raw = pred.get(key, {})
            pred_val = (
                pred_raw.get("value", "") if isinstance(pred_raw, dict)
                else str(pred_raw)
            ).upper().strip()

            # Get ground truth value — may be flat string or nested
            gt_raw = gt.get(key, "")
            gt_val = (
                gt_raw.get("value", "") if isinstance(gt_raw, dict)
                else str(gt_raw)
            ).upper().strip()

            totals[key] += 1
            if pred_val == gt_val:
                matches[key] += 1

    per_column = {
        key: round(matches[key] / totals[key], 3) if totals[key] else 0.0
        for key in FIELD_KEYS
    }
    overall = round(sum(per_column.values()) / len(per_column), 3)

    return {
        "per_column": per_column,
        "overall": overall,
        "records_compared": n,
        "column_labels": {
            k: EXCEL_COLUMN_NAMES[i]
            for i, k in enumerate(FIELD_KEYS)
        },
    }


# Evaluate from uploaded Excel file 

@app.post("/evaluate-file")
async def evaluate_from_file(
    ground_truth_file: UploadFile = File(...),
    predictions: str = Body(...),
):
    """
    Accept ground truth as an uploaded Excel/CSV file.
    Accept predictions as a JSON string.
    Returns per-column accuracy scores.
    """
    import openpyxl
    import csv

    pred_records = json.loads(predictions)

    content = await ground_truth_file.read()
    filename = ground_truth_file.filename or ""

    gt_records = []

    if filename.endswith(".csv"):
        decoded = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(decoded))
        for row in reader:
            gt_records.append(_gt_row_to_dict(row))

    elif filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value).strip() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, [str(v).strip() if v else "" for v in row]))
            gt_records.append(_gt_row_to_dict(row_dict))
    else:
        raise HTTPException(400, "Only .xlsx and .csv ground truth files are supported")

    if not gt_records:
        raise HTTPException(400, "Ground truth file is empty or could not be parsed")

    # 1. Create Ground Truth Dictionary keyed by barcode
    gt_map = {}
    for gt in gt_records:
        barcode = str(gt.get("barcode", "")).upper().strip()
        if barcode:
            gt_map[barcode] = gt

    totals  = {k: 0 for k in FIELD_KEYS}
    matches = {k: 0 for k in FIELD_KEYS}
    records_compared = 0
    unmatched_predictions = 0

    # 2. Iterate through predictions and lookup by barcode
    for pred in pred_records:
        pred_barcode_raw = pred.get("barcode", {})
        pred_barcode = (
            pred_barcode_raw.get("value", "") if isinstance(pred_barcode_raw, dict)
            else str(pred_barcode_raw)
        ).upper().strip()

        if not pred_barcode or pred_barcode not in gt_map:
            unmatched_predictions += 1
            continue

        gt = gt_map[pred_barcode]
        records_compared += 1

        for key in FIELD_KEYS:
            pred_raw = pred.get(key, {})
            pred_val = (
                pred_raw.get("value", "") if isinstance(pred_raw, dict)
                else str(pred_raw)
            ).upper().strip()

            gt_val = str(gt.get(key, "")).upper().strip()

            totals[key] += 1
            if pred_val == gt_val:
                matches[key] += 1

    per_column = {
        key: round(matches[key] / totals[key], 3) if totals[key] else 0.0
        for key in FIELD_KEYS
    }
    overall = round(sum(per_column.values()) / len(per_column), 3) if per_column else 0.0

    return {
        "per_column": per_column,
        "overall": overall,
        "records_compared": records_compared,
        "unmatched_predictions": unmatched_predictions,
        "gt_total": len(gt_records),
        "pred_total": len(pred_records),
        "column_labels": {
            k: EXCEL_COLUMN_NAMES[i]
            for i, k in enumerate(FIELD_KEYS)
        },
    }


def _gt_row_to_dict(row: dict) -> dict:
    """
    Map ground truth Excel column names to our internal field keys.
    Handles the exact column names from the hackathon dataset.
    """
    mapping = {
        "ITEM_NAME":         "item_name",
        "BARCODE":           "barcode",
        "MANUFACTURER":      "manufacturer",
        "BRAND":             "brand",
        "WEIGHT":            "weight",
        "PACKAGING TYPE":    "packaging_type",
        "PACKAGING_TYPE":    "packaging_type",
        "COUNTRY":           "country",
        "VARIANT":           "variant",
        "TYPE":              "type",
        "FRAGRANCE_FLAVOR":  "fragrance_flavor",
        "PROMOTION":         "promotion",
        "ADDONS":            "addons",
        "TAGLINE":           "tagline",
    }
    result = {}
    for excel_col, field_key in mapping.items():
        result[field_key] = row.get(excel_col, "")
    return result