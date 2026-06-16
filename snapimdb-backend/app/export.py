import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from app.schemas import FIELD_KEYS, EXCEL_COLUMN_NAMES


# Color palettes 

# Header
HEADER_BG    = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")

# Confidence fills
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")  # ≥ 0.75
YELLOW_FILL  = PatternFill("solid", fgColor="FFEB9C")  # 0.50 – 0.74
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")  # < 0.50
EMPTY_FILL   = PatternFill("solid", fgColor="F2F2F2")  # empty value

# Duplicate row highlight
DUP_FILL     = PatternFill("solid", fgColor="FCE4D6")  # light orange

# Conflict note row
CONFLICT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow

# Fonts
DATA_FONT    = Font(size=9, name="Calibri")
CONFLICT_FONT = Font(size=8, name="Calibri", italic=True, color="7F6000")
DUP_FONT     = Font(size=9, name="Calibri", color="833C00")

# Border
THIN_BORDER  = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Column widths (index matches FIELD_KEYS order)
COLUMN_WIDTHS = [
    52,  # ITEM_NAME
    18,  # BARCODE
    28,  # MANUFACTURER
    20,  # BRAND
    12,  # WEIGHT
    18,  # PACKAGING TYPE
    18,  # COUNTRY
    16,  # VARIANT
    20,  # TYPE
    18,  # FRAGRANCE_FLAVOR
    20,  # PROMOTION
    20,  # ADDONS
    28,  # TAGLINE
]


def _get_confidence_fill(confidence: float, value: str) -> PatternFill:
    if not value:
        return EMPTY_FILL
    if confidence >= 0.75:
        return GREEN_FILL
    if confidence >= 0.50:
        return YELLOW_FILL
    return RED_FILL


def _get_field(record: dict, field: str) -> tuple[str, float]:
    """Returns (value, confidence) for a field in a record dict."""
    raw = record.get(field, {})
    if isinstance(raw, dict):
        return raw.get("value", ""), float(raw.get("confidence", 0.0))
    return str(raw) if raw else "", 0.0


def build_predictions_excel(records: list[dict]) -> bytes:
    """
    Build a predictions.xlsx file from a list of IMDBRecord dicts.

    Features:
    - Confidence color coding per cell (green/yellow/red)
    - Duplicate rows highlighted in orange with a note
    - Conflict rows get a sub-row with the conflict note
    - Frozen header row
    - Auto column widths
    - Summary stats sheet

    Returns raw bytes ready to stream as a file download.
    """
    wb = Workbook()

    # Main predictions sheet 
    ws = wb.active
    ws.title = "Predictions"

    # Header row
    for col_idx, col_name in enumerate(EXCEL_COLUMN_NAMES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = HEADER_BG
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_idx - 1]

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"  # Freeze header

    current_row = 2

    for record in records:
        is_duplicate = record.get("is_duplicate", False)
        has_conflicts = record.get("has_conflicts", False)
        conflict_notes = record.get("conflict_notes", "")

        # Data row 
        for col_idx, field_key in enumerate(FIELD_KEYS, start=1):
            value, confidence = _get_field(record, field_key)

            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=False
            )

            if is_duplicate:
                cell.fill = DUP_FILL
                cell.font = DUP_FONT
            else:
                cell.fill = _get_confidence_fill(confidence, value)
                cell.font = DATA_FONT

        ws.row_dimensions[current_row].height = 16
        current_row += 1

        #  Duplicate note row 
        if is_duplicate:
            dup_of = record.get("duplicate_of", "unknown")
            note_cell = ws.cell(
                row=current_row,
                column=1,
                value=f"⚠ DUPLICATE — matches product: {dup_of}"
            )
            note_cell.fill  = DUP_FILL
            note_cell.font  = Font(size=8, italic=True, color="833C00", name="Calibri")
            note_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(EXCEL_COLUMN_NAMES)
            )
            ws.row_dimensions[current_row].height = 13
            current_row += 1

        # Conflict note row 
        if has_conflicts and conflict_notes:
            conf_cell = ws.cell(
                row=current_row,
                column=1,
                value=f" CRITIC NOTE: {conflict_notes}"
            )
            conf_cell.fill  = CONFLICT_FILL
            conf_cell.font  = CONFLICT_FONT
            conf_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(EXCEL_COLUMN_NAMES)
            )
            ws.row_dimensions[current_row].height = 13
            current_row += 1

    # Summary stats sheet 
    ws2 = wb.create_sheet(title="Summary")

    total = len(records)
    duplicates = sum(1 for r in records if r.get("is_duplicate", False))
    conflicts  = sum(1 for r in records if r.get("has_conflicts", False))

    # Per-column confidence averages
    col_stats = {}
    for field_key in FIELD_KEYS:
        confs = []
        for rec in records:
            _, conf = _get_field(rec, field_key)
            val, _ = _get_field(rec, field_key)
            if val:
                confs.append(conf)
        col_stats[field_key] = round(sum(confs) / len(confs), 3) if confs else 0.0

    overall_conf = round(
        sum(col_stats.values()) / len(col_stats), 3
    ) if col_stats else 0.0

    # Summary header
    ws2["A1"] = "SnapIMDB — Extraction Summary"
    ws2["A1"].font = Font(bold=True, size=13, name="Calibri", color="1E3A5F")
    ws2.merge_cells("A1:C1")

    summary_rows = [
        ("", "", ""),
        ("Metric", "Value", ""),
        ("Total products extracted", total, ""),
        ("Duplicate records flagged", duplicates, ""),
        ("Records with critic conflicts", conflicts, ""),
        ("Overall avg confidence", f"{round(overall_conf * 100, 1)}%", ""),
        ("", "", ""),
        ("Field", "Avg Confidence", "Status"),
    ]

    for row_data in summary_rows:
        ws2.append(list(row_data))

    # Per-column confidence rows
    for field_key, conf in col_stats.items():
        col_label = EXCEL_COLUMN_NAMES[FIELD_KEYS.index(field_key)]
        pct = round(conf * 100, 1)
        status = "Good" if conf >= 0.75 else "⚠ Review" if conf >= 0.5 else "Weak"
        row = ws2.max_row + 1
        ws2.cell(row=row, column=1, value=col_label)
        ws2.cell(row=row, column=2, value=f"{pct}%")
        ws2.cell(row=row, column=3, value=status)

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 15

    # Stream to bytes 
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()